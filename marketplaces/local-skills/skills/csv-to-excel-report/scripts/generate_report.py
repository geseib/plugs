#!/usr/bin/env python3
"""
Generate a styled Excel report from sales data and targets CSVs.

Usage: python3 generate_report.py <sales.csv> <targets.csv> <output.xlsx>
"""

import argparse
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------

DARK_BLUE = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
LIGHT_BLUE = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
LIGHT_GRAY = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
GREEN_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
RED_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")

HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=16, color="1F4E79")
SUBTITLE_FONT = Font(bold=False, size=11, color="666666")
KPI_LABEL_FONT = Font(bold=True, size=10, color="1F4E79")
KPI_VALUE_FONT = Font(bold=True, size=14, color="1F4E79")
TABLE_HEADER_FONT = Font(bold=True, size=10, color="1F4E79")
BOLD_FONT = Font(bold=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

CURRENCY_FMT = '$#,##0.00'
PCT_FMT = '0.0%'
INT_FMT = '#,##0'
DATE_FMT = 'YYYY-MM-DD'

SALES_COLUMNS = ["date", "region", "sales_rep", "product_category",
                 "units_sold", "unit_price", "revenue"]
TARGET_COLUMNS = ["region", "sales_rep", "quarter", "target_revenue"]


# ---------------------------------------------------------------------------
# Step 1: Validate CSVs
# ---------------------------------------------------------------------------

def validate_csv(path, required_columns):
    """Read a CSV and verify it has the required columns."""
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"File not found: {path}")

    df = pd.read_csv(p)
    df.columns = df.columns.str.strip().str.lower()
    found = set(df.columns)
    required = set(required_columns)
    missing = required - found

    if missing:
        extra = found - required
        suggestions = []
        for m in sorted(missing):
            for e in extra:
                if m.replace("_", "") == e.replace("_", "") or m in e or e in m:
                    suggestions.append(f"  '{e}' -> '{m}'?")
        msg = f"Missing columns in {p.name}: {sorted(missing)}\nFound: {sorted(found)}"
        if suggestions:
            msg += "\nPossible renames:\n" + "\n".join(suggestions)
        raise ValueError(msg)

    return df


# ---------------------------------------------------------------------------
# Step 3: Load and merge
# ---------------------------------------------------------------------------

def derive_quarter(date_series):
    """Convert date strings to quarter format like '2024-Q1'."""
    dt = pd.to_datetime(date_series)
    return dt.dt.year.astype(str) + "-Q" + dt.dt.quarter.astype(str)


def merge_data(sales_df, targets_df):
    """Left-join sales onto targets, calculate metrics."""
    sales_df["quarter"] = derive_quarter(sales_df["date"])

    merged = sales_df.merge(
        targets_df,
        on=["region", "sales_rep", "quarter"],
        how="left",
    )

    merged["achievement_pct"] = merged.apply(
        lambda r: r["revenue"] / r["target_revenue"]
        if pd.notna(r["target_revenue"]) and r["target_revenue"] > 0
        else None,
        axis=1,
    )
    merged["variance"] = merged.apply(
        lambda r: r["revenue"] - r["target_revenue"]
        if pd.notna(r["target_revenue"])
        else None,
        axis=1,
    )

    return merged


# ---------------------------------------------------------------------------
# Step 4: Calculate KPIs
# ---------------------------------------------------------------------------

def calculate_kpis(merged_df):
    """Compute executive summary KPIs."""
    total_revenue = merged_df["revenue"].sum()
    total_units = merged_df["units_sold"].sum()
    num_transactions = len(merged_df)
    avg_deal_size = total_revenue / num_transactions if num_transactions else 0
    num_reps = merged_df["sales_rep"].nunique()

    regional = (
        merged_df.groupby("region")
        .agg(
            revenue=("revenue", "sum"),
            target=("target_revenue", "sum"),
            transactions=("revenue", "count"),
        )
        .reset_index()
    )
    regional["achievement_pct"] = regional.apply(
        lambda r: r["revenue"] / r["target"] if pd.notna(r["target"]) and r["target"] > 0 else None,
        axis=1,
    )
    regional["variance"] = regional.apply(
        lambda r: r["revenue"] - r["target"] if pd.notna(r["target"]) else None,
        axis=1,
    )
    regional = regional.sort_values("revenue", ascending=False)

    top5 = (
        merged_df.groupby("sales_rep")
        .agg(revenue=("revenue", "sum"), deals=("revenue", "count"))
        .reset_index()
        .sort_values("revenue", ascending=False)
        .head(5)
    )

    return {
        "total_revenue": total_revenue,
        "total_units": int(total_units),
        "avg_deal_size": avg_deal_size,
        "num_reps": num_reps,
        "num_transactions": num_transactions,
        "regional": regional,
        "top5": top5,
        "date_min": merged_df["date"].min(),
        "date_max": merged_df["date"].max(),
    }


# ---------------------------------------------------------------------------
# Step 5: Executive Summary sheet
# ---------------------------------------------------------------------------

def _apply_cell(ws, row, col, value, font=None, fill=None, fmt=None, alignment=None, border=None):
    """Helper to set a cell's value and style."""
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    return cell


def build_executive_summary(ws, kpis):
    """Populate the Executive Summary sheet."""
    ws.sheet_properties.tabColor = "1F4E79"

    # Title area
    _apply_cell(ws, 1, 1, "Regional Sales Performance Report", font=TITLE_FONT)
    ws.merge_cells("A1:F1")
    _apply_cell(ws, 2, 1, f"Period: {kpis['date_min']} to {kpis['date_max']}", font=SUBTITLE_FONT)
    ws.merge_cells("A2:F2")

    # KPI cards (row 4-5, 2x2 layout)
    kpi_items = [
        ("Total Revenue", kpis["total_revenue"], CURRENCY_FMT),
        ("Total Units Sold", kpis["total_units"], INT_FMT),
        ("Avg Deal Size", kpis["avg_deal_size"], CURRENCY_FMT),
        ("Active Sales Reps", kpis["num_reps"], INT_FMT),
    ]
    positions = [(4, 1), (4, 4), (6, 1), (6, 4)]

    for (label, value, fmt), (row, col) in zip(kpi_items, positions):
        _apply_cell(ws, row, col, label, font=KPI_LABEL_FONT, fill=LIGHT_BLUE, border=THIN_BORDER)
        _apply_cell(ws, row, col + 1, value, font=KPI_VALUE_FONT, fill=LIGHT_BLUE,
                     fmt=fmt, border=THIN_BORDER)
        # Extend fill to col+2 for visual balance
        _apply_cell(ws, row, col + 2, None, fill=LIGHT_BLUE, border=THIN_BORDER)

    # Regional Breakdown (starting row 9)
    row = 9
    _apply_cell(ws, row, 1, "Regional Breakdown", font=Font(bold=True, size=12, color="1F4E79"))
    ws.merge_cells(f"A{row}:F{row}")
    row += 1

    headers = ["Region", "Revenue", "Target", "Achievement %", "Variance", "Transactions"]
    for c, h in enumerate(headers, 1):
        _apply_cell(ws, row, c, h, font=HEADER_FONT, fill=DARK_BLUE,
                     alignment=Alignment(horizontal="center"), border=THIN_BORDER)
    row += 1

    regional = kpis["regional"]
    for _, rdata in regional.iterrows():
        _apply_cell(ws, row, 1, rdata["region"], font=BOLD_FONT, border=THIN_BORDER)
        _apply_cell(ws, row, 2, rdata["revenue"], fmt=CURRENCY_FMT, border=THIN_BORDER)
        _apply_cell(ws, row, 3, rdata["target"] if pd.notna(rdata["target"]) else "N/A",
                     fmt=CURRENCY_FMT if pd.notna(rdata["target"]) else None, border=THIN_BORDER)
        pct = rdata["achievement_pct"]
        _apply_cell(ws, row, 4, pct if pd.notna(pct) else "N/A",
                     fmt=PCT_FMT if pd.notna(pct) else None, border=THIN_BORDER)
        var = rdata["variance"]
        cell = _apply_cell(ws, row, 5, var if pd.notna(var) else "N/A",
                            fmt=CURRENCY_FMT if pd.notna(var) else None, border=THIN_BORDER)
        if pd.notna(var):
            cell.fill = GREEN_FILL if var >= 0 else RED_FILL
        _apply_cell(ws, row, 6, int(rdata["transactions"]), fmt=INT_FMT, border=THIN_BORDER)
        row += 1

    # Top 5 Performers (after regional table)
    row += 1
    _apply_cell(ws, row, 1, "Top 5 Performers", font=Font(bold=True, size=12, color="1F4E79"))
    ws.merge_cells(f"A{row}:D{row}")
    row += 1

    top_headers = ["Rank", "Sales Rep", "Revenue", "Deals"]
    for c, h in enumerate(top_headers, 1):
        _apply_cell(ws, row, c, h, font=HEADER_FONT, fill=DARK_BLUE,
                     alignment=Alignment(horizontal="center"), border=THIN_BORDER)
    row += 1

    for rank, (_, rep) in enumerate(kpis["top5"].iterrows(), 1):
        _apply_cell(ws, row, 1, rank, font=BOLD_FONT, border=THIN_BORDER,
                     alignment=Alignment(horizontal="center"))
        _apply_cell(ws, row, 2, rep["sales_rep"], border=THIN_BORDER)
        _apply_cell(ws, row, 3, rep["revenue"], fmt=CURRENCY_FMT, border=THIN_BORDER)
        _apply_cell(ws, row, 4, int(rep["deals"]), fmt=INT_FMT, border=THIN_BORDER)
        row += 1

    # Auto-size columns
    for col_idx in range(1, 7):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18


# ---------------------------------------------------------------------------
# Step 6: Sales Data sheet
# ---------------------------------------------------------------------------

def build_data_sheet(ws, merged_df):
    """Populate the filterable data sheet."""
    ws.sheet_properties.tabColor = "4472C4"

    # Prepare display columns
    display_cols = ["date", "region", "sales_rep", "product_category",
                    "units_sold", "unit_price", "revenue", "quarter",
                    "target_revenue", "achievement_pct", "variance"]
    display_headers = ["Date", "Region", "Sales Rep", "Product Category",
                       "Units Sold", "Unit Price", "Revenue", "Quarter",
                       "Target Revenue", "Achievement %", "Variance"]

    # Column format map
    col_formats = {
        "unit_price": CURRENCY_FMT,
        "revenue": CURRENCY_FMT,
        "target_revenue": CURRENCY_FMT,
        "variance": CURRENCY_FMT,
        "achievement_pct": PCT_FMT,
        "units_sold": INT_FMT,
    }

    # Write headers
    for c, header in enumerate(display_headers, 1):
        _apply_cell(ws, 1, c, header, font=HEADER_FONT, fill=DARK_BLUE,
                     alignment=Alignment(horizontal="center"), border=THIN_BORDER)

    # Write data rows
    for r_idx, (_, row_data) in enumerate(merged_df.iterrows(), 2):
        for c_idx, col in enumerate(display_cols, 1):
            val = row_data.get(col)
            if pd.isna(val):
                val = ""
            fmt = col_formats.get(col)
            fill = LIGHT_GRAY if r_idx % 2 == 0 else WHITE_FILL
            _apply_cell(ws, r_idx, c_idx, val, fmt=fmt, fill=fill, border=THIN_BORDER)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filters
    last_col = get_column_letter(len(display_headers))
    last_row = len(merged_df) + 1
    ws.auto_filter.ref = f"A1:{last_col}{last_row}"

    # Auto-size columns
    col_widths = {"date": 12, "region": 14, "sales_rep": 18, "product_category": 18,
                  "units_sold": 12, "unit_price": 13, "revenue": 14, "quarter": 11,
                  "target_revenue": 16, "achievement_pct": 15, "variance": 14}
    for c_idx, col in enumerate(display_cols, 1):
        ws.column_dimensions[get_column_letter(c_idx)].width = col_widths.get(col, 14)


# ---------------------------------------------------------------------------
# Step 7: Validate output
# ---------------------------------------------------------------------------

def validate_output(path, expected_data_rows):
    """Open the generated file and verify structure."""
    wb = load_workbook(path, read_only=True)
    sheets = wb.sheetnames
    assert "Executive Summary" in sheets, "Missing 'Executive Summary' sheet"
    assert "Sales Data" in sheets, "Missing 'Sales Data' sheet"

    data_ws = wb["Sales Data"]
    actual_rows = data_ws.max_row - 1  # minus header
    wb.close()

    if actual_rows != expected_data_rows:
        print(f"  WARNING: Expected {expected_data_rows} data rows, found {actual_rows}")
    else:
        print(f"  Data rows: {actual_rows} (matches expected)")

    return True


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Generate styled Excel report from CSV data.")
    parser.add_argument("sales_csv", help="Path to sales data CSV")
    parser.add_argument("targets_csv", help="Path to sales targets CSV")
    parser.add_argument("output_xlsx", help="Path for output Excel file")
    args = parser.parse_args()

    # Step 1: Validate
    print("Step 1: Validating input CSVs...")
    try:
        sales_df = validate_csv(args.sales_csv, SALES_COLUMNS)
        targets_df = validate_csv(args.targets_csv, TARGET_COLUMNS)
    except (ValueError, FileNotFoundError) as e:
        print(f"ERROR: {e}", file=sys.stderr)
        sys.exit(1)
    print(f"  Sales data: {len(sales_df)} rows")
    print(f"  Targets: {len(targets_df)} rows")

    # Step 3: Merge
    print("Step 3: Merging data...")
    merged = merge_data(sales_df, targets_df)
    missing_targets = merged["target_revenue"].isna().sum()
    if missing_targets:
        print(f"  WARNING: {missing_targets} rows have no matching target")
    print(f"  Merged dataset: {len(merged)} rows")

    # Step 4: KPIs
    print("Step 4: Calculating KPIs...")
    kpis = calculate_kpis(merged)
    print(f"  Total revenue: ${kpis['total_revenue']:,.2f}")
    print(f"  Active reps: {kpis['num_reps']}")

    # Step 5-6: Build workbook
    print("Step 5-6: Generating Excel workbook...")
    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "Executive Summary"
    build_executive_summary(summary_ws, kpis)

    data_ws = wb.create_sheet("Sales Data")
    build_data_sheet(data_ws, merged)

    output_path = Path(args.output_xlsx)
    wb.save(output_path)
    print(f"  Saved: {output_path}")

    # Step 7: Validate
    print("Step 7: Validating output...")
    validate_output(output_path, len(merged))

    print(f"\nDone. Report: {output_path}")
    print(f"  Sheets: Executive Summary, Sales Data")
    print(f"  Rows: {len(merged)}")
    if missing_targets:
        print(f"  Note: {missing_targets} transactions had no matching quarterly target")


if __name__ == "__main__":
    main()
